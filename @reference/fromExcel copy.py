from openpyxl import load_workbook
from openpyxl import utils
import dbf

testRange = 6
file = 'IO_INSTRUMENT_LIST.xlsx'
sheetPrototypes = 'DEFAULTS'
sheetTags = 'TAG_REF'
tagKey = 'TAGNAME'
divider = '____________________________________________________________________'

wb = load_workbook(filename=file, read_only=False,
                   keep_vba=False, data_only=True, keep_links=False)
ws = wb[sheetPrototypes]
ws2 = wb[sheetTags]


def main(fileName):
    data = getData(fileName)
    createDBF(data)


def getData(fileName):
    print(divider)
    print('Getting data from sheet')

    data = {
        'file': fileName,
        'sheet': sheetPrototypes,
        'protoNames': [],
        'fieldNames': [],
        'entries': [],
        'variables': {},
        'tags': {},
        'refColumns': [],

    }
    # subFunction definintions -------------------------------------------------

    def populateFieldNames():
        for rowIndex in range(3, ws.max_row+1):
            cell = ws['A'+str(rowIndex)]
            if cell.value != None:
                data['fieldNames'].append(cell.value)
        print('Got FieldNames')

    def populateProtoNames():
        for colIndex in range(2, ws.max_column+1):
            cell = ws[utils.get_column_letter(colIndex)+'2']
            if cell.value != None:
                data['protoNames'].append(cell.value)
        print('Got ProtoNames')

    def populateEntries():
        for colIndex in range(2, ws.max_column+1):
            rows = []
            for rowIndex in range(3, ws.max_row+1):
                cell = ws[utils.get_column_letter(colIndex)+str(rowIndex)]
                rows.append(cell.value)
            data['entries'].append(tuple(rows))
        data['entries'] = tuple(data['entries'])
        print('Got Entries')

    # def populateVariables():
    #     for colIndex in range(2, ws.max_column+1):
    #         rows = []
    #         for rowIndex in range(3, ws.max_row+1):
    #             cell = ws[utils.get_column_letter(colIndex)+str(rowIndex)]
    #             if iterable(cell.value):
    #                 if '##' in cell.value:
    #                     rows.append(str(ctVariable(cextraell.value)))
    #         data['variables'].append(rows)
    #     print('Got Variables')

    def populateTags():
        for protoName in data['protoNames']:

            data['tags'][protoName] = []
            for rowIndex in range(3, 19):
                tagName = ws2['B'+str(rowIndex)].value
                if protoName is ws2['D'+str(rowIndex)].value:
                    data['tags'][protoName].append(tagName)

    def populateRefColumns():
        for columnIndex in range(2, ws.max_column+1):
            cell = ws2[utils.get_column_letter(columnIndex)+'2']
            if cell.value != None:
                data['refColumns'].append(cell.value)

    # --------------------------------------------------------------------------
    populateFieldNames()
    populateProtoNames()
    populateEntries()

    populateTags()
    populateRefColumns()

    # updateVariableEntries(0)
    # addTagData(tagData)
    # addVariables()

    # populateVariables()
    print('')
    print('Got All Data')
    print(divider)
    return data


def updateVariableEntries(data, index):
    # print(index)

    def getTagData(index, refColumn):
        index += 3
        if refColumn in data['refColumns']:
            columnLetter = utils.get_column_letter(
                data['refColumns'].index(refColumn)+2)
            cell = ws2[columnLetter+str(index)]
            return cell.value

    data['entries'] = []
    for colIndex in range(2, ws.max_column+1):
        rows = []
        for rowIndex in range(3, ws.max_row+1):
            cell = ws[utils.get_column_letter(colIndex)+str(rowIndex)]
            if iterable(cell.value):
                if '##' in cell.value:
                    variable = extractVariable(cell.value)
                    tagDatum = getTagData(index, variable)
                    if tagDatum != None:
                        rows.append(tagDatum)
                        print('Added '+tagDatum)
                    else:
                        rows.append(cell.value)
                    # rows.append(getTagData(0,refColumn))
                else:
                    rows.append(cell.value)

            else:
                rows.append(cell.value)
        data['entries'].append(tuple(rows))
    data['entries'] = tuple(data['entries'])


def createDBF(data):
    print(divider)
    print('Generating * .dbfs')
    print('')

    # subFunction definintions -------------------------------------------------
    def createTable(protoName, protoIndex, fieldStr):
        if fieldsStr != '':
            dbfTable = dbf.Table(filename='outputDBFs/'+protoName +
                                 '.dbf', field_specs=fieldsStr,)
            dbfTable.open(mode=dbf.READ_WRITE)
            # for
            tableRow = tuple([str(e) for e in data['entries']
                              [protoIndex] if e != None])
            # dbfTable.append(tableRow)
            for index, tag in enumerate(data['tags'][protoName]):
                # print(index, tag)
                updateVariableEntries(data, index)
                dbfTable.append(tableRow)
                # print(index, tag)
                # print(data['entries'])
            print(protoName+'.dbf created successfully')
        else:
            print(protoName + " is empty, no file created")

    def createFieldsStr(protoIndex):
        output = ''
        for fieldName in data['fieldNames']:
            fieldIndex = data['fieldNames'].index(fieldName)
            if data['entries'][protoIndex][fieldIndex] != None:
                output += fieldName + ' C(255); '
        return output

    # --------------------------------------------------------------------------
    for protoName in data['protoNames']:
        protoIndex = data['protoNames'].index(protoName)
        fieldsStr = createFieldsStr(protoIndex)
        createTable(protoName, protoIndex, fieldsStr)

    print('')
    print('DBFs created')
    print(divider)

# Helper Functions -------------------------------------------------------------


def iterable(obj):
    try:
        iter(obj)
    except Exception:
        return False
    else:
        return True


def extractVariable(inputString):
    x = inputString
    return x[x.find('##')+2: inputString.rfind('##')]


# ------------------------------------------------------------------------------
# ------------------------------------------------------------------------------
main(file)
