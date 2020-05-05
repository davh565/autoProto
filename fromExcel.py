from openpyxl import load_workbook
from openpyxl import utils
import dbf

testRange = 6
file = 'IO_INSTRUMENT_LIST.xlsx'
sheet = 'DEFAULTS'
divider = '____________________________________________________________________'


def main(fileName, sheetName):
    data = getData(fileName, sheetName)
    # print(data['protoNames'])
    # print(data['fieldNames'])
    createDBF(data)


def getData(fileName, sheetName):
    print(divider)
    print('Getting data from sheet')

    wb = load_workbook(filename=file, read_only=False,
                       keep_vba=False, data_only=True, keep_links=False)
    ws = wb[sheetName]

    data = {
        'file': fileName,
        'sheet': sheetName,
        'protoNames': [],
        'fieldNames': [],
        'entries': [],
        'variables': []

    }
    # subFunction definintions -------------------------------------------------

    def getFieldNames():
        for rowIndex in range(3, ws.max_row+1):
            cell = ws['A'+str(rowIndex)]
            if cell.value != None:
                data['fieldNames'].append(cell.value)
        print('Got FieldNames')

    def getProtoNames():
        for colIndex in range(2, ws.max_column+1):
            cell = ws[utils.get_column_letter(colIndex)+'2']
            if cell.value != None:
                data['protoNames'].append(cell.value)
        print('Got ProtoNames')

    def getEntries():
        for colIndex in range(2, ws.max_column+1):
            rows = []
            for rowIndex in range(3, ws.max_row+1):
                cell = ws[utils.get_column_letter(colIndex)+str(rowIndex)]
                rows.append(cell.value)
            data['entries'].append(tuple(rows))
        data['entries'] = tuple(data['entries'])
        print('Got Entries')

    # def getVariables():
    #     for colIndex in range(2, testRange):
    #         rows = []
    #         for rowIndex in range(3, testRange):
    #             cell = ws[utils.get_column_letter(colIndex)+str(rowIndex)]
    #             if iterable(cell.value):
    #                 if '##' in cell.value:
    #                     rows.append(str(cell.value))
    #                 else:
    #                     rows.append(None)
    #             else:
    #                 rows.append(None)
    #         data['variables'].append(rows)
    #     print('Got Variables')
    # --------------------------------------------------------------------------
    getFieldNames()
    getProtoNames()
    getEntries()
    # getVariables()
    print('')
    print('Got All Data')
    print(divider)
    return data


def createDBF(data):
    print(divider)
    print('Generating * .dbfs')
    print('')

    # subFunction definintions -------------------------------------------------
    def createTable(protoName, protoIndex, fieldStr):
        if fieldsStr != '':
            dbfTable = dbf.Table(filename=protoName +
                                 '.dbf', field_specs=fieldsStr,)
            dbfTable.open(mode=dbf.READ_WRITE)
            entries = tuple([str(e) for e in data['entries']
                             [protoIndex] if e != None])
            dbfTable.append(entries)
            print(protoName+'.dbf created successfully')
        else:
            print(protoName + " is empty, no file created")

    def createFieldsStr(protoIndex):
        output = ''
        for fieldName in data['fieldNames']:
            fieldIndex = data['fieldNames'].index(fieldName)
            if data['entries'][protoIndex][fieldIndex] != None:
                output += fieldName + ' C(255); '
        return output

    # --------------------------------------------------------------------------
    for protoName in data['protoNames']:
        protoIndex = data['protoNames'].index(protoName)
        fieldsStr = createFieldsStr(protoIndex)
        createTable(protoName, protoIndex, fieldsStr)

    print('')
    print('DBFs created')
    print(divider)

# Helper Functions -------------------------------------------------------------


def iterable(obj):
    try:
        iter(obj)
    except Exception:
        return False
    else:
        return True


# ------------------------------------------------------------------------------
# ------------------------------------------------------------------------------
main(file, sheet)
