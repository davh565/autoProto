from openpyxl import load_workbook
from openpyxl import utils


def getData(params):
    print('''
________________________________________________________________________________
Getting Data:
''')
    p = params
    wb = load_workbook(filename='input/'+p['inputFile'], read_only=False,
                       keep_vba=False, data_only=True, keep_links=False)
    ws = wb[p['sheetPrototypes']]
    ws2 = wb[p['sheetTags']]

    data = {
        'protoNames': [],
        'fieldNames': [],
        'protoFieldValues': [],
        'tagNames': [],
        'dataNames': [],
        'tagDataValues': [],
    }
    # subFunction definintions -------------------------------------------------

    def populateProtoNames():
        for colIndex in range(2, ws.max_column+1):
            cell = ws[utils.get_column_letter(colIndex)+'2']
            if cell.value != None:
                data['protoNames'].append(cell.value)
        print('Got ProtoNames')

    def populateFieldNames():
        for rowIndex in range(3, ws.max_row+1):
            cell = ws['A'+str(rowIndex)]
            if cell.value != None:
                data['fieldNames'].append(cell.value)
        print('Got FieldNames')

    def populateProtoFieldValues():
        for colIndex in range(2, ws.max_column+1):
            rows = []
            for rowIndex in range(3, ws.max_row+1):
                cell = ws[utils.get_column_letter(colIndex)+str(rowIndex)]
                rows.append(cell.value)
            data['protoFieldValues'].append(rows)
        print('Got protoFieldValues')

    def populateDataNames():
        for colIndex in range(2, ws2.max_column+1):
            cell = ws2[utils.get_column_letter(colIndex)+'2']
            if cell.value != None:
                data['dataNames'].append(cell.value)
        print('Got dataNames')

    def populateTagNames():
        for rowIndex in range(3, ws2.max_row+1):
            cell = ws2['A'+str(rowIndex)]
            if cell.value != None:
                data['tagNames'].append(cell.value)
        print('Got tagNames')

    def populateTagDataValues():
        for colIndex in range(2, ws2.max_column+1):
            rows = []
            for rowIndex in range(3, ws2.max_row+1):
                cell = ws2[utils.get_column_letter(colIndex)+str(rowIndex)]
                rows.append(cell.value)
            data['tagDataValues'].append(rows)
        print('Got tagDataValues')
    # ---------------------------------------------------------------------------
    populateProtoNames()
    populateFieldNames()
    populateProtoFieldValues()
    populateTagNames()
    populateDataNames()
    populateTagDataValues()
    print('''
Got all data
--------------------------------------------------------------------------------
''')
    return data
