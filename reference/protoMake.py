# protoMake.py
# By David Humphries
# v0.1-29.03.19
#
# Asks a series of questions via console prompt regarding prototype configuration. Generates a set of headers
# based on responses in .xlsx & .dbf format.
#
# Todo:
#  * Make all existing headers compliant with v5 prototype spec
#  * Add support for additional items from v5 spec (motors, vsd, soft starters etc.)
#  * Refactor to clean up excessive repeated code
#  * Error handling
#  * Extend capabilities to generate data in addition to column headers
#
############


from openpyxl import Workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import dbf

########################################################################


def question(queryString, answerType):
    answer = input(queryString)
    if(answerType == "number"):
        if(answer.isdigit()):
            return int(answer)
        elif(answer == "n" or answer == "N" or answer == "0"):
            answer = 0
            return int(answer)
        else:
            print("Please enter a valid number, or 'N' to skip.")
            answer = question(queryString, answerType)
            return int(answer)

    elif(answerType == "boolean"):
        if(answer == "y" or answer == "Y" or answer == "1"):
            answer = 1
            return answer
        elif(answer == "n" or answer == "N" or answer == "0"):
            answer = 0
            return answer
        else:
            print("Please enter a valid boolean. (y/n or 1/0)")
            answer = question(queryString, answerType)
    else:
        return answer
########################################################################


########################################################################
def protoMake():
    # Prototype Name
    ####################################
    prototype = question("Enter Protoype Name: ", "string")

    # Panel Setup
    ####################################
    panels = question("Enter Number of Panels: ", "number")
    if(panels):
        power24DC = []
        power24AC = []
        power230AC = []
        xIO = []
        xIOF = []
        xIOK = []
        xDB = []
        xDBF = []
        xDBK = []
        for n in range(panels):
            power24DC.append(
                question("Is there 24VDC power in Panel " + str(n+1) + "? (Y/N): ", "boolean"))
            power24AC.append(
                question("Is there 24VAC power in Panel " + str(n+1) + "? (Y/N): ", "boolean"))
            power230AC.append(
                question("How many 230VAC phases in Panel " + str(n+1) + "?: ", "number"))

            xIO.append(question("Are there IO terminals in Panel " +
                                str(n+1) + "? (Y/N): ", "boolean"))
            if(xIO[n]):
                xIOF.append(question("Are there IO fuses in Panel " +
                                     str(n+1) + "? (Y/N): ", "boolean"))
                xIOK.append(
                    question("Are there IO relays in Panel " + str(n+1) + "? (Y/N): ", "boolean"))
            else:
                xIOF.append(0)
                xIOK.append(0)
            xDB.append(question(
                "Are there distribution terminals in Panel " + str(n+1) + "? (Y/N): ", "boolean"))
            if(xDB[n]):
                xDBF.append(question(
                    "Are there distribution fuses or MCBs in Panel " + str(n+1) + "? (Y/N): ", "boolean"))
                xDBK.append(question(
                    "Are there distribution relays or contactors in Panel " + str(n+1) + "? (Y/N): ", "boolean"))
            else:
                xDBF.append(0)
                xDBK.append(0)

    # PLC Setup
    ####################################
    plcs = question("Enter Number of PLCs: ", "number")
    if(plcs):
        plcT = []
        for n in range(plcs):
            plcT.append(
                question("Enter Number of terminals for PLC " + str(n+1) + ": ", "number"))
        print(plcT)

    # Device Setup
    ####################################
    devices = question("Enter Number of Devices: ", "number")
    if(devices):
        deviceT = []
        for n in range(devices):
            deviceT.append(
                question("Enter Number of terminals for Device " + str(n+1) + ": ", "number"))
        print(plcT)

    # Cable Setup
    ####################################
    cables = question("Enter Number of Cables: ", "number")
    if(cables):
        cores = []
        for n in range(cables):
            cores.append(
                question("Enter Number of cores for Cable " + str(n+1) + ": ", "number"))
        print(cores)

    #########################################################################
    headers = ["DRAWING", "PROTOTYPE"]

    ####################################
    # PANELS
    ####################################
    if(panels):
        for n in range(panels):
            prefix = "X"+str(n+1)+"_"
            headers.append(prefix+"PNL")
            headers.append(prefix+"PDESC")

            ####################################
            # POWER
            ####################################
            if(power24DC[n]):
                headers.append(prefix+"24DC")
                headers.append(prefix+"0DC")
            if(power24AC[n]):
                headers.append(prefix+"24ACL")
                headers.append(prefix+"24ACN")
            for m in range(power230AC[n]):
                headers.append(prefix+"230ACL"+str(m+1))
            if(power230AC[n]):
                headers.append(prefix+"230ACN")
            ####################################
            # IO
            ####################################
            if(xIO[n]):
                headers.append(prefix+"XIO")
                headers.append(prefix+"XIO_TI")
            if(xIOF[n]):
                headers.append(prefix+"XIO_FI")
            if(xIOK[n]):
                headers.append(prefix+"XIO_KI")

            ####################################
            # DISTRIBUTION
            ####################################
            if(xDB[n]):
                headers.append(prefix+"XDB")
                headers.append(prefix+"XDB_TI")
            if(xDBF[n]):
                headers.append(prefix+"XDB_FI")
            if(xDBK[n]):
                headers.append(prefix+"XDB_KI")

    ####################################
    # PLCS
    ####################################
    if(plcs):
        for n in range(plcs):
            prefix = "P"+str(n+1)+"_"
            headers.append(prefix+"NAME")
            headers.append(prefix+"RACK")
            headers.append(prefix+"SLOT")
            headers.append(prefix+"CHAN")
            headers.append(prefix+"MFG")
            headers.append(prefix+"CATN")
            headers.append(prefix+"CATD")
            headers.append(prefix+"TAG")
            headers.append(prefix+"DESC2")
            headers.append(prefix+"DESC3")
            headers.append(prefix+"DESC4")
            for m in range(plcT[n]):
                headers.append(prefix+"T"+str(m+1))

    ####################################
    # DEVICES
    ####################################
    if(devices):
        for n in range(devices):
            prefix = "D"+str(n+1)+"_"
            headers.append(prefix+"TAG")
            headers.append(prefix+"DESC2")
            headers.append(prefix+"DESC3")
            headers.append(prefix+"DESC4")
            headers.append(prefix+"VDESC")
            headers.append(prefix+"TYPE")
            for m in range(deviceT[n]):
                headers.append(prefix+"T"+str(m+1))

    ####################################
    # CABLES
    ####################################
    if(cables):
        for n in range(cables):
            prefix = "C"+str(n+1)+"_"
            headers.append(prefix+"MFG")
            headers.append(prefix+"CATN")
            headers.append(prefix+"CATD")
            headers.append(prefix+"TAG")
            for m in range(cores[n]):
                headers.append(prefix+"COR"+str(m+1))

    ####################################
    # OUTPUT
    ####################################
    print("Generating " + prototype + ".xlsx")

    wb = Workbook()

    ws = wb.active
    ws.append(headers)
    tab = Table(displayName="Table1", ref="A1:" +
                get_column_letter(ws.max_column)+"2")
    style = TableStyleInfo(name="TableStyleMedium13", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(prototype + '.xlsx')

    ####################################
    print("Generating " + prototype + ".dbf")

    specs = ""
    for item in headers:
        specs += item + " C(255); "
    dbfTable = dbf.Table(
        filename=prototype,
        field_specs=specs,
    )
    dbfTable.open(mode=dbf.READ_WRITE)

    print(headers)
    ########################################################################
    # REPEAT
    ####################################
    if(question("Would you like to make another Prototype? (Y/N): ", "boolean")):
        protoMake()


protoMake()
